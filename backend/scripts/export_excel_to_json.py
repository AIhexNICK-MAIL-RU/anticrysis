#!/usr/bin/env python3
"""
Экспорт Excel-книги в JSON: все листы (включая скрытые), ячейки, формулы.
Запуск из корня проекта:
  python backend/scripts/export_excel_to_json.py "Рабочая тетрадь_Фин модель.xlsx"
  python backend/scripts/export_excel_to_json.py  # ищет файл в текущей директории
Результат: backend/app/calc_model/exported/
"""
from __future__ import annotations

import json
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)


def safe_filename(name: str) -> str:
    """Имя листа → безопасное имя файла."""
    return re.sub(r'[^\w\- ]+', '_', name).strip() or "sheet"


def export_workbook(path: str, out_dir: str) -> None:
    wb = load_workbook(path, read_only=False, data_only=False)
    os.makedirs(out_dir, exist_ok=True)

    book_meta: dict = {
        "sheets": [],
        "active_sheet": wb.active.title if wb.active else "",
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data: dict = {
            "name": sheet_name,
            "visibility": getattr(ws, "sheet_state", "visible") or "visible",
            "dimensions": ws.calculate_dimension() or "",
            "cells": {},
        }
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = cell.value
                if cell.data_type == "f":
                    sheet_data["cells"][cell.coordinate] = {
                        "value": str(val),
                        "formula": str(val),
                        "data_type": "f",
                    }
                else:
                    sheet_data["cells"][cell.coordinate] = {
                        "value": val if isinstance(val, (int, float)) else str(val),
                        "formula": None,
                        "data_type": cell.data_type or "n",
                    }
        fname = safe_filename(sheet_name) + ".json"
        out_path = os.path.join(out_dir, fname)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(sheet_data, f, ensure_ascii=False, indent=2)
        book_meta["sheets"].append({
            "name": sheet_name,
            "file": fname,
            "visibility": sheet_data["visibility"],
        })
        print(f"  {sheet_name} -> {fname}")

    with open(os.path.join(out_dir, "_workbook.json"), "w", encoding="utf-8") as f:
        json.dump(book_meta, f, ensure_ascii=False, indent=2)
    print(f"Экспортировано {len(wb.sheetnames)} листов в {out_dir}")


def main() -> None:
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    if len(sys.argv) >= 2:
        xlsx_path = os.path.abspath(sys.argv[1])
    else:
        xlsx_path = os.path.join(root, "Рабочая тетрадь_Фин модель.xlsx")
    if not os.path.isfile(xlsx_path):
        print(f"Файл не найден: {xlsx_path}")
        print("Использование: python export_excel_to_json.py [путь/к/файлу.xlsx]")
        sys.exit(1)
    out_dir = os.path.join(root, "backend", "app", "calc_model", "exported")
    export_workbook(xlsx_path, out_dir)


if __name__ == "__main__":
    main()
